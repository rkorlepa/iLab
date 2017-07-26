#!/ws/rkorlepa-sjc/python/bin/python

import os, sys, time
import pexpect
import re
import logging
import traceback
import collections
import tablib
from prompt_toolkit import prompt
from subprocess import call
from openpyxl import load_workbook

from libilab.Database import Switches

'''Used for clear console only'''
SWITCH_PROMPT = '#'
CONSOLE_PROMPT = '[a-zA-Z0-9_-]+[>|#]'
SWITCH_LOGIN = '(?i)login:'
PWD_PROMPT = '(?i)password:'
LOGIN_INCORRECT = '(?i)Login incorrect'
PX2_POWER_PROMPT = 'PX2'
PX_POWER_PROMPT = 'clp:/->'
APC_POWER_PROMPT = 'User Name'
LICENSE_PROMPT = 'understand this risk and wish to continue'

console_pwd_list = ['nbv123','cisco123','nbv_1234']
switch_pwd_list = ['nbv_12345','nbv_1234', 'cisco_123', 'Rtp13579']

result_dict = collections.defaultdict(dict)

def loginIlabDB(ip, log):
    pwd = str(Switches.get(Switches.mgmt_ip==ip).switch_pwd)
    logging.info("Try loginIlabDB with %s[%s]", ip, pwd)
    console = pexpect.spawn('telnet -l admin %s' % (ip))
    console.logfile = log
    i = console.expect([pexpect.TIMEOUT, pexpect.EOF, LOGIN_INCORRECT, \
            PWD_PROMPT, SWITCH_PROMPT], 5)
    if i==0 or i==1:
        logging.info('loginIlabDB Unknown IP')
    if i==2:
        logging.info("loginIlabDB pwd %s in ilab is not correct", pwd)
    if i==3:
        console.sendline(pwd)
        j = console.expect([pexpect.TIMEOUT, pexpect.EOF, LOGIN_INCORRECT, \
                SWITCH_LOGIN, PWD_PROMPT, LICENSE_PROMPT, SWITCH_PROMPT], 5)
        while j>=0:
            if j==0 or j==1:
                logging.info("loginIlabDB Weird console")
                break
            if j==2 or j==3 or j==4:
                logging.info("loginIlabDB pwd %s in ilab is not correct", pwd)
                break
            if j==5 or j==6:
                logging.info("loginIlabDB pwd[%s] match", pwd)
                result_dict[ip]['login'] = pwd
                result_dict[ip]['ilab_pwd'] = pwd
                #query = ipdistinguish.update(login=pwd).where(ipdistinguish.ip==ip)
                #ret = query.execute()
                console.close()
                return True
            j = console.expect([pexpect.TIMEOUT, pexpect.EOF, SWITCH_LOGIN, PWD_PROMPT, LICENSE_PROMPT, SWITCH_PROMPT], 5)
    if i==4:
        logging.info("loginIlabDB no pwd required")
        result_dict[ip]['login'] = 'N/A'
        result_dict[ip]['ilab_pwd'] = pwd
        #query = ipdistinguish.update(login='N/A').where(ipdistinguish.ip==ip)
        #ret = query.execute()
        console.close()
        return True

    console.close()
    result_dict[ip]['ilab_pwd'] = pwd
    return False

def login3Attempts(ip, log):
    temp_list = switch_pwd_list[:]
    ilab_pwd = str(Switches.get(Switches.mgmt_ip==ip).switch_pwd)
    if ilab_pwd in temp_list:
        temp_list.remove(ilab_pwd)

    console = pexpect.spawn('telnet %s' % (ip))
    console.logfile = log
    console.sendline('')
    k = 0
    i = console.expect([pexpect.TIMEOUT, pexpect.EOF, SWITCH_LOGIN], 5)
    if i==0 or i==1:
        logging.info('login3Attempts Unknown IP')
    if i==2:
        logging.info('login3Attempts Found a Switch')
        console.sendline('admin')
        console.expect(PWD_PROMPT)
        console.sendline(temp_list[k])
        j = console.expect([pexpect.TIMEOUT, pexpect.EOF, SWITCH_LOGIN, PWD_PROMPT, LICENSE_PROMPT, SWITCH_PROMPT], 5)
        while j>=0:
            if j==0 or j==1:
                logging.info("login3Attempts Weird console")
                break
            if j==2:
                logging.info("login3Attempts Invalid pwd [%s] try another one", \
                        temp_list[k])
                console.sendline('admin')
                k = k + 1
                if k > 3:
                    break
                pass
            if j==3:
                console.sendline(temp_list[k])
                pass
            if j==4 or j==5:
                result_dict[ip]['login'] = temp_list[k] 
                #query = ipdistinguish.update(login=temp_list[k]).where(ipdistinguish.ip==ip)
                #ret = query.execute()
                break
            j = console.expect([pexpect.TIMEOUT, pexpect.EOF, SWITCH_LOGIN, PWD_PROMPT, LICENSE_PROMPT, SWITCH_PROMPT], 5)

    console.close()
    return

def what_is_the_ip(ip, log):
    logging.info('Telnet to console with ip %s',ip)
    console = pexpect.spawn('telnet %s' % (ip))
    console.logfile = log
    k = 0
    i = console.expect([pexpect.TIMEOUT, pexpect.EOF, SWITCH_LOGIN, \
        PWD_PROMPT, PX_POWER_PROMPT, PX2_POWER_PROMPT, APC_POWER_PROMPT], 10)
    if i==0 or i==1:
        console.sendline('')
        m = console.expect([pexpect.TIMEOUT, pexpect.EOF, SWITCH_LOGIN], 5)
        if m==0 or m==1:
            logging.info('Unknown IP')
            result_dict[ip] = {'device':'Unknown','login':''}
            #ipdistinguish.create(ip=ip,device='Unknown')
        if m==2:
            logging.infor('Found a Switch')
            result_dict[ip] = {'device':'Switch','login':''}
            #ipdistinguish.create(ip=ip,device='Switch')
    if i==2:
        logging.info('Found a Switch')
        result_dict[ip] = {'device':'Switch','login':''}
        #ipdistinguish.create(ip=ip,device='Switch')
    if i==3:
        console.sendline(console_pwd_list[k])
        j = console.expect([pexpect.TIMEOUT, pexpect.EOF, PWD_PROMPT, CONSOLE_PROMPT, 'Bad'], 5)
        while j >= 0:
            if j==0 or j==1:
                logging.info("Weird console")
                result_dict[ip] = {'device':'UnknownConsole','login':''}
                #ipdistinguish.create(ip=ip,device='UnknownConsole')
                break
            if j==2:
                logging.info("Invalid pwd [%s] try another one", \
                        console_pwd_list[k])
                k = k + 1
                if k > 3:
                    break
                pass
            if j==3:
                console.sendline('terminal length 0')
                console.expect(CONSOLE_PROMPT)
                console.sendline('show line | include TTY')
                console.expect(CONSOLE_PROMPT)
                temp = console.before
                if temp:
                    logging.info("Found a Console")
                    result_dict[ip] = {'device':'Console','login':''}
                    #ipdistinguish.create(ip=ip,device='Console')
                break
            if j==4:
                logging.info('wrong pwd for console')
                result_dict[ip] = {'device':'Console','login':''}
                #ipdistinguish.create(ip=ip,device='Console')
                break
            console.sendline(console_pwd_list[k])
            j = console.expect([pexpect.TIMEOUT, pexpect.EOF, PWD_PROMPT, CONSOLE_PROMPT, 'Bad'], 5)
    if i==4 or i==5 or i==6:
        logging.info("Found Power Slot")
        result_dict[ip] = {'device':'PowerSlot','login':''}
        #ipdistinguish.create(ip=ip,device='PowerSlot')

    console.close()
    '''
    temp_ip = ipdistinguish.get(ipdistinguish.ip==ip)
    if str(temp_ip.device).lower() == 'switch':
        logging.info("Try to get the password of switch")
        success = loginIlabDB(ip,log)
        if not(success):
            login3Attempts(ip,log)
    '''
    if result_dict[ip]['device'].lower() == 'switch':
        logging.info("Try to get the password of switch")
        success = loginIlabDB(ip,log)
        if not(success):
            login3Attempts(ip,log)
    return

def create_xls_file(u_mail):
    xls_databook = tablib.Databook()
    xls_data_sheet = []
    headers_sheet = ('IP', 'Device', 'Login Pwd', 'iLab Pwd')

    for ip,detail in result_dict.iteritems():
        temp_data_sheet = (str(ip),
                           str(detail['device']),
                           str(detail['login']),
                           str[detail['ilab_pwd'])
        xls_data_sheet.append(temp_data_sheet)

    xls_data_sheet = tablib.Dataset(*xls_data_sheet, headers=headers_sheet, \
            title='IP Details')
    xls_databook.add_sheet(xls_data_sheet)

    with open('ip_details.xls', 'wb') as f:
        f.write(xls_databook.xls)

    xls_template = ilab_template("xlsmail.txt").render(user=str(u_mail))
    u_mail = str(u_mail) + '@cisco.com'
    mailer = IlabEmail('rkorlepa@cisco.com', u_mail)
    mailer.set_subject("iLab[IP Details]")
    mailer.add_line(xls_template)
    file_path = os.getcwd() + '/ip_details.xls'
    mailer.set_attachment(file_path)
    mailer.send()
    mailer.reset()

if __name__ == '__main__':
    u_mail = prompt(u'Mail to whom(CEC username): ')

    strfile = 'ipDistng.log'
    call(["rm", strfile])
    fout = None
    fout = file(strfile, 'a')
    logging.basicConfig(filename=strfile, format='%(asctime)s: %(levelname)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', level=logging.DEBUG)

    print "Sit back and enjoy you will get a mail at %s@cisco.com" % u_mail
    print "Do Not Hit CTRL+C" % u_mail
    wb = load_workbook('Fretta_IPs.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(1,sheet.max_row):
        m = re.search('\d+.\d+.\d+.\d+', str(sheet.cell(row=i,column=1).value))
        if m:
            ips = str(m.group(0))
            what_is_the_ip(ips, fout)

    # Send mail to u_mail = cec username
    create_xls_file(u_mail)
